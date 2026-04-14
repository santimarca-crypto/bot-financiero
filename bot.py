
import os, re, sqlite3, logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BOT_TOKEN = os.environ["BOT_TOKEN"]
DB = Path("/data/ops.db")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# compro/vendo nombre 1300 x 1365  (x con o sin espacios)
OP_RE = re.compile(
    r'\b(compro|compra|vendo|venta)\b\s+'
    r'([\w][\w\s]*?)\s+'
    r'([\d][\d.,]*)'
    r'\s*[xXaA@]\s*([\d][\d.,]*)',
    re.IGNORECASE
)

# compro/vendo [nombre] 1000000/1380  (pesos dividido TC)
OP_RE2 = re.compile(
    r'\b(compro|compra|vendo|venta)\b\s+'
    r'(?:([a-zA-Z][\w ]*?)\s+)?'
    r'([\d][\d.,]+)\s*/\s*([\d][\d.,]+)',
    re.IGNORECASE
)

# salen/entran 500000 [nombre opcional]
MOV_RE = re.compile(
    r'\b(salen|entran)\b\s+([\d][\d.,]+)(?:\s+([a-zA-Z][\w\s]*))?',
    re.IGNORECASE
)

def num(s):
    s = s.strip().replace(" ","")
    if s.count(".") == 1 and len(s.split(".")[1]) == 3:
        s = s.replace(".", "")
    elif s.count(",") == 1 and len(s.split(",")[1]) == 3:
        s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return float(s.replace(".","").replace(",",""))

def fmt(n):
    return "{:,.0f}".format(abs(n)).replace(",",".")

def fmtd(n):
    return "{:,.2f}".format(abs(n)).replace(",","X").replace(".","!").replace("X",".").replace("!",",")

def db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def setup():
    DB.parent.mkdir(parents=True, exist_ok=True)
    with db() as c:
        c.execute("CREATE TABLE IF NOT EXISTS cfg (k TEXT PRIMARY KEY, v REAL DEFAULT 0)")
        c.execute("CREATE TABLE IF NOT EXISTS ops (id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT, hora TEXT, de TEXT, tipo TEXT, contra TEXT, usd REAL, tc REAL, ars REAL, msg TEXT)")
        c.execute("INSERT OR IGNORE INTO cfg VALUES ('usd',0)")
        c.execute("INSERT OR IGNORE INTO cfg VALUES ('ars',0)")
        c.commit()

def cfg(k):
    with db() as c:
        r = c.execute("SELECT v FROM cfg WHERE k=?", (k,)).fetchone()
        return r["v"] if r else 0.0

def setcfg(k, v):
    with db() as c:
        c.execute("INSERT OR REPLACE INTO cfg VALUES (?,?)", (k,v))
        c.commit()

def guardar(de, tipo, contra, usd, tc, ars, msg):
    now = datetime.now()
    with db() as c:
        c.execute("INSERT INTO ops (fecha,hora,de,tipo,contra,usd,tc,ars,msg) VALUES (?,?,?,?,?,?,?,?,?)",
                  (now.strftime("%d/%m/%Y"), now.strftime("%H:%M:%S"), de, tipo, contra, usd, tc, ars, msg))
        c.commit()

def num_op_hoy():
    hoy = datetime.now().strftime("%d/%m/%Y")
    with db() as c:
        r = c.execute("SELECT COUNT(*) FROM ops WHERE fecha=?", (hoy,)).fetchone()
        return r[0]

def posicion():
    pu, pa = cfg("usd"), cfg("ars")
    with db() as c:
        for r in c.execute("SELECT tipo,usd,ars FROM ops ORDER BY id").fetchall():
            if r["tipo"] == "Compra":
                pu += r["usd"]; pa -= r["ars"]
            elif r["tipo"] == "Venta":
                pu -= r["usd"]; pa += r["ars"]
            elif r["tipo"] == "Salida":
                pa -= r["ars"]
            elif r["tipo"] == "Entrada":
                pa += r["ars"]
    return pu, pa

async def start(u: Update, _):
    await u.message.reply_text(
        "Bot USD/ARS activo\n\n"
        "Operaciones con tipo de cambio:\n"
        "compro Richard 1300 x 1365\n"
        "vendo Vicky Kantai 5000 x 1382\n\n"
        "Operaciones con pesos/TC:\n"
        "compro 1000000/1380\n"
        "vendo Richard 900000/1465\n\n"
        "Movimientos de pesos:\n"
        "salen 500000\n"
        "entran 200000\n\n"
        "/posicion - ver posicion de caja\n"
        "/historial - operaciones de hoy\n"
        "/excel - bajar Excel\n"
        "/inicio ARS - fijar saldo inicial en pesos\n"
        "/corregir ID USD TC - corregir una operacion\n"
        "/cancelar ID - cancelar una operacion\n"
        "/resetear - borrar todo y arrancar de cero"
    )

async def pos_cmd(u: Update, _):
    _, pa = posicion()
    sa = "+" if pa >= 0 else "-"
    await u.message.reply_text(
        "POSICION DE CAJA\n"
        "ARS: " + sa + fmt(pa)
    )

async def hist_cmd(u: Update, ctx):
    hoy = datetime.now().strftime("%d/%m/%Y")
    with db() as c:
        rows = c.execute("SELECT * FROM ops WHERE fecha=? ORDER BY id", (hoy,)).fetchall()
    if not rows:
        await u.message.reply_text("No hay operaciones hoy.")
        return
    txt = "Operaciones de hoy " + hoy + ":\n\n"
    for i, r in enumerate(rows, 1):
        if r["tipo"] in ("Salida", "Entrada"):
            nombre_h = " " + r["contra"] if r["contra"] != "-" else ""
            txt += "#" + str(i) + " " + r["hora"][:5] + " | " + r["tipo"].upper() + nombre_h + " $" + fmt(r["ars"]) + " (ID:" + str(r["id"]) + ")\n"
        else:
            e = "COMPRA" if r["tipo"]=="Compra" else "VENTA"
            txt += "#" + str(i) + " " + r["hora"][:5] + " | " + e + " " + r["contra"] + " USD " + fmtd(r["usd"]) + " x " + fmt(r["tc"]) + " (ID:" + str(r["id"]) + ")\n"
    await u.message.reply_text(txt)

async def inicio_cmd(u: Update, ctx):
    if len(ctx.args) < 1:
        await u.message.reply_text("Uso: /inicio ARS\nEj: /inicio -500000")
        return
    try:
        ars_i = num(ctx.args[0])
        setcfg("ars", ars_i)
        signo = "-" if ars_i < 0 else ""
        await u.message.reply_text("Saldo inicial en pesos: " + signo + "$" + fmt(ars_i))
    except Exception as e:
        await u.message.reply_text("Error: " + str(e))

async def cancelar_cmd(u: Update, ctx):
    if not ctx.args:
        await u.message.reply_text("Uso: /cancelar ID\nEj: /cancelar 25\n(el ID lo ves en /historial entre parentesis)")
        return
    op_id = ctx.args[0]
    with db() as c:
        r = c.execute("SELECT * FROM ops WHERE id=?", (op_id,)).fetchone()
        if not r:
            await u.message.reply_text("No existe operacion con ID " + str(op_id))
            return
        e = r["tipo"].upper()
        c.execute("DELETE FROM ops WHERE id=?", (op_id,))
        c.commit()
    _, pa = posicion()
    if r["tipo"] in ("Salida", "Entrada"):
        detalle = e + " $" + fmt(r["ars"])
    else:
        detalle = e + " " + r["contra"] + " USD " + fmtd(r["usd"]) + " x " + fmt(r["tc"])
    await u.message.reply_text(
        "CANCELADA operacion ID " + str(op_id) + "\n"
        + detalle + "\n\n"
        "POSICION DE CAJA\n"
        "ARS: " + ("+" if pa>=0 else "-") + fmt(pa)
    )

async def corregir_cmd(u: Update, ctx):
    if len(ctx.args) < 3:
        await u.message.reply_text("Uso: /corregir ID USD TC\nEj: /corregir 25 2000 1380\n(el ID lo ves en /historial entre parentesis)")
        return
    try:
        op_id = int(ctx.args[0])
        usd_v = num(ctx.args[1])
        tc_v = num(ctx.args[2])
        ars_v = usd_v * tc_v
        with db() as c:
            r = c.execute("SELECT * FROM ops WHERE id=?", (op_id,)).fetchone()
            if not r:
                await u.message.reply_text("No existe operacion con ID " + str(op_id))
                return
            c.execute("UPDATE ops SET usd=?, tc=?, ars=? WHERE id=?", (usd_v, tc_v, ars_v, op_id))
            c.commit()
        _, pa = posicion()
        await u.message.reply_text(
            "CORREGIDA operacion ID " + str(op_id) + "\n"
            "USD " + fmtd(usd_v) + " x $ " + fmt(tc_v) + "\n\n"
            "POSICION DE CAJA\n"
            "ARS: " + ("+" if pa>=0 else "-") + fmt(pa)
        )
    except Exception as e:
        await u.message.reply_text("Error: " + str(e))

async def resetear_cmd(u: Update, ctx):
    with db() as c:
        c.execute("DELETE FROM ops")
        c.execute("DELETE FROM sqlite_sequence WHERE name='ops'")
        c.execute("UPDATE cfg SET v=0 WHERE k='ars'")
        c.commit()
    await u.message.reply_text(
        "RESETEO COMPLETO\n"
        "Todas las operaciones borradas.\n"
        "Saldo en cero.\n\n"
        "Ahora usa /inicio para fijar el saldo de apertura."
    )

async def excel_cmd(u: Update, _):
    await u.message.reply_text("Generando Excel...")
    with db() as c:
        rows = c.execute("SELECT * FROM ops ORDER BY id").fetchall()
    wb = Workbook(); ws = wb.active; ws.title = "Operaciones"
    hdrs = ["#","Fecha","Hora","Enviado por","Tipo","Contraparte","USD","TC","ARS","Pos ARS"]
    navy = "1F4E79"
    def tb():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s,right=s,top=s,bottom=s)
    hf = PatternFill("solid", start_color=navy)
    for c2, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c2, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = hf; cell.border = tb()
        cell.alignment = Alignment(horizontal="center")
    DS = 2
    ars_i = cfg("ars")
    for i, r in enumerate(rows):
        row = DS + i
        alt = PatternFill("solid", start_color="EBF3FB" if i%2==0 else "FFFFFF")
        b = tb()
        vals = [r["id"], r["fecha"], r["hora"], r["de"], r["tipo"], r["contra"], r["usd"], r["tc"]]
        for c2, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c2, value=v)
            cell.fill = alt; cell.border = b
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=9, value=r["ars"]).number_format = "#,##0.00"
        ws.cell(row=row, column=9).fill = alt; ws.cell(row=row, column=9).border = b
        if row == DS:
            if r["tipo"] == "Compra":
                pa = "="+str(ars_i)+"-I"+str(row)
            elif r["tipo"] == "Venta":
                pa = "="+str(ars_i)+"+I"+str(row)
            elif r["tipo"] == "Salida":
                pa = "="+str(ars_i)+"-I"+str(row)
            else:
                pa = "="+str(ars_i)+"+I"+str(row)
        else:
            if r["tipo"] == "Compra":
                pa = "=J"+str(row-1)+"-I"+str(row)
            elif r["tipo"] == "Venta":
                pa = "=J"+str(row-1)+"+I"+str(row)
            elif r["tipo"] == "Salida":
                pa = "=J"+str(row-1)+"-I"+str(row)
            else:
                pa = "=J"+str(row-1)+"+I"+str(row)
        cell = ws.cell(row=row, column=10, value=pa)
        cell.fill = alt; cell.border = b; cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="center")
    for c2, w in enumerate([4,12,9,18,10,14,12,12,14,13], 1):
        ws.column_dimensions[get_column_letter(c2)].width = w
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = "operaciones_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx"
    await u.message.reply_document(document=buf.read(), filename=fname)

async def mensaje(u: Update, ctx):
    try:
        text = u.message.text
        sender = u.effective_user.first_name or u.effective_user.username or "?"
        log.info("MSG de " + sender + ": " + repr(text))

        # Movimientos de pesos: salen/entran
        mov = MOV_RE.search(text)
        if mov:
            kw, monto_s, nombre_s = mov.groups()
            monto = num(monto_s)
            tipo = "Salida" if kw.lower() == "salen" else "Entrada"
            contra = nombre_s.strip().title() if nombre_s else "-"
            guardar(sender, tipo, contra, 0, 0, monto, text)
            n_hoy = num_op_hoy()
            _, pa = posicion()
            sp = "-" if tipo == "Salida" else "+"
            nombre_txt = " " + contra if contra != "-" else ""
            resp = (
                "OK " + tipo.upper() + " #" + str(n_hoy) + "\n"
                + nombre_txt.strip() + (" | " if contra != "-" else "") + sp + "$ " + fmt(monto) + "\n\n"
                "POSICION DE CAJA\n"
                "ARS: " + ("+" if pa>=0 else "-") + fmt(pa)
            )
            await u.message.reply_text(resp)
            return

        # Operacion con pesos/TC
        m2 = OP_RE2.search(text)
        if m2 and "/" in text:
            kw, contra_s, pesos_s, tc_s = m2.groups()
            pesos = num(pesos_s)
            tc_v = num(tc_s)
            usd_v = pesos / tc_v
            contra = (contra_s.strip().title() if contra_s else "-")
            tipo = "Compra" if kw.lower() in ("compro","compra") else "Venta"
            guardar(sender, tipo, contra, usd_v, tc_v, pesos, text)
            n_hoy = num_op_hoy()
            _, pa = posicion()
            sp = "+" if tipo == "Compra" else "-"
            sp2 = "-" if tipo == "Compra" else "+"
            resp = (
                ("OK COMPRA" if tipo=="Compra" else "OK VENTA") + " #" + str(n_hoy) + "\n"
                + contra + " | USD " + fmtd(usd_v) + " x $ " + fmt(tc_v) + "\n"
                + sp2 + "$ " + fmt(pesos) + "\n\n"
                "POSICION DE CAJA\n"
                "ARS: " + ("+" if pa>=0 else "-") + fmt(pa)
            )
            await u.message.reply_text(resp)
            return

        # Operacion con USD x TC
        m = OP_RE.search(text)
        if m:
            kw, contra, usd_s, tc_s = m.groups()
            usd_v = num(usd_s)
            contra = contra.strip().title()
            tc_v = num(tc_s)
            tipo = "Compra" if kw.lower() in ("compro","compra") else "Venta"
            guardar(sender, tipo, contra, usd_v, tc_v, usd_v*tc_v, text)
            n_hoy = num_op_hoy()
            _, pa = posicion()
            sp2 = "-" if tipo == "Compra" else "+"
            ars_op = usd_v * tc_v
            resp = (
                ("OK COMPRA" if tipo=="Compra" else "OK VENTA") + " #" + str(n_hoy) + "\n"
                + contra + " | USD " + fmtd(usd_v) + " x $ " + fmt(tc_v) + "\n"
                + sp2 + "$ " + fmt(ars_op) + "\n\n"
                "POSICION DE CAJA\n"
                "ARS: " + ("+" if pa>=0 else "-") + fmt(pa)
            )
            await u.message.reply_text(resp)
            return

        log.info("Sin match")

    except Exception as e:
        log.error("ERROR: " + str(e), exc_info=True)
        try:
            await u.message.reply_text("Error: " + str(e))
        except:
            pass

def main():
    setup()
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("posicion", pos_cmd))
    app.add_handler(CommandHandler("historial", hist_cmd))
    app.add_handler(CommandHandler("excel", excel_cmd))
    app.add_handler(CommandHandler("inicio", inicio_cmd))
    app.add_handler(CommandHandler("cancelar", cancelar_cmd))
    app.add_handler(CommandHandler("corregir", corregir_cmd))
    app.add_handler(CommandHandler("resetear", resetear_cmd))
    app.add_handler(MessageHandler(filters.ALL, mensaje))
    log.info("Bot iniciado")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
